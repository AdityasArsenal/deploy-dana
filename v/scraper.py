import requests
import os
from openpyxl import load_workbook

def download_xml_files(file_path, limit) -> None:
    os.makedirs('new_xml', exist_ok=True)
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    wb = load_workbook(file_path)
    ws = wb.active
    
    company_names = []
    urls = []
    file_names = []

    j=1
    for i in range(1,len(ws["A"])): 

        company_names.append((ws["A"][i].value))
        urls.append((ws["E"][i].value))
        j+=1


    for i, url in enumerate(urls, 0):
        if i > limit:                                             
          break
        try:
            filename = os.path.join('new_xml', f"{company_names[i]}.xml")
            file_names.append(filename)

            print(f"Downloading file {i} of {limit}...")
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            # Save in new directory 
            with open(filename, "wb") as f:
                f.write(response.content)
            
        except requests.RequestException as e:
            print(f"❌ Error downloading {url}: {str(e)}")
    
    # Remove any duplicate company names while preserving order
    seen = set()
    unique_companies = []
    unique_files = []
    unique_urls = []

    for i, (company, file, url) in enumerate(zip(company_names, file_names, urls)):
        if company not in seen:
            seen.add(company)
            unique_companies.append(company)
            unique_files.append(file) 
            unique_urls.append(url)

    company_names = unique_companies
    file_names = unique_files
    urls = unique_urls


    return company_names,file_names,urls

