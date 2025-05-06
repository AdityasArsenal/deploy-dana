import requests
import os
from openpyxl import load_workbook

def download_xml_files(file_path, folder_name, limit) -> None:
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    wb = load_workbook(file_path)
    ws = wb.active
    
    company_names = []
    xml_urls = []
    brsr_urls = []
    xml_file_names = []
    brsr_file_names = []

    for i in range(1,len(ws["A"])): 
        company_names.append((ws["A"][i].value))
        xml_urls.append((ws["E"][i].value))
        brsr_urls.append((ws["D"][i].value))

    for i, (xml_url, brsr_url) in enumerate(zip(xml_urls, brsr_urls), 0):
        if i == limit:                                             
          break
        try:
            # Create company-specific folder
            company_folder = os.path.join(f'{folder_name}', company_names[i])
            os.makedirs(company_folder, exist_ok=True)
            
            xml_filename = os.path.join(company_folder, f"{company_names[i]}.xml")
            brsr_filename = os.path.join(company_folder, f"{company_names[i]}.pdf")
            xml_file_names.append(xml_filename)
            brsr_file_names.append(brsr_filename)

            print(f"Downloading files for {company_names[i]} ({i+1} of {limit})...")
            xml_response = requests.get(xml_url, headers=headers)
            brsr_response = requests.get(brsr_url, headers=headers)     

            xml_response.raise_for_status()
            brsr_response.raise_for_status()
            
            # Save in company folder
            with open(xml_filename, "wb") as f:
                f.write(xml_response.content)
            with open(brsr_filename, "wb") as f:
                f.write(brsr_response.content)

            print(f"Downloaded files for {company_names[i]} ({i+1} of {limit})...")

        except requests.RequestException as e:
            print(f"❌ Error downloading files for {company_names[i]}: {str(e)}")
    
    return company_names
