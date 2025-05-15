import requests
import os
from openpyxl import load_workbook
from file_to_blob import upload_folder_to_blob
import time


def download_files(excel_file_path, destination_folder_name, parent_folder_name, limit, blob_container_name, container_client) -> None:
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    all_uploaded_blobs = []
    urls_of_files_uploaded = [] #to save the urls of the files uploaded to blob storage

    wb = load_workbook(excel_file_path)
    ws = wb.active
    
    company_names = []
    xml_urls = []
    brsr_urls = []
    xml_file_names = []
    brsr_file_names = []
    failed_to_download_files = []

    for i in range(0,len(ws["A"])): 
        company_names.append((ws["A"][i].value))
        xml_urls.append((ws["E"][i].value))
        brsr_urls.append((ws["D"][i].value))

    for i, (xml_url, brsr_url) in enumerate(zip(xml_urls, brsr_urls), 0):

        if i == limit:                                             
          break 
 
        try:
            # Create company-specific folder
            company_folder = os.path.join(f'{destination_folder_name}', company_names[i])
            os.makedirs(company_folder, exist_ok=True)
            
            # Create company-specific xml and brsr files
            xml_filename = os.path.join(company_folder, f"{company_names[i]}.xml")
            brsr_filename = os.path.join(company_folder, f"{company_names[i]}.pdf")
            xml_file_names.append(xml_filename)
            brsr_file_names.append(brsr_filename)

            # Download xml and brsr files
            print(f"Downloading files for {company_names[i]} ({i+1} of {limit})...")
            xml_response = requests.get(xml_url, headers=headers)
            brsr_response = requests.get(brsr_url, headers=headers)     

            xml_response.raise_for_status()
            brsr_response.raise_for_status()
            
            # Save in company specific files
            with open(xml_filename, "wb") as f:
                f.write(xml_response.content)
            with open(brsr_filename, "wb") as f:
                f.write(brsr_response.content)

            print(f"Downloaded files for {company_names[i]} ({i+1} of {limit})...")

            # if i % 10 == 0 and i != 0:
            if i == 9:
            #upload the files to blob storage after 10 files are downloaded and remove the files from the local folder after uploading to blob storage

                files_removed = 0

                for company in os.listdir(parent_folder_name):
                    company_folder = os.path.join(parent_folder_name, company)

                    urls_uploaded_of_this_company, blobs_uploaded_of_this_company = upload_folder_to_blob(blob_container_name, company_folder, container_client)

                    urls_of_files_uploaded.extend(urls_uploaded_of_this_company)
                    all_uploaded_blobs.extend(blobs_uploaded_of_this_company)

                    print(f"Uploaded file from {company} to blob storage /n URLs is : {urls_uploaded_of_this_company}/n")

                    # remove the files from the local folder after uploading to blob storage
                    max_retries = 3
                    retry_delay = 0.5  # seconds

                    for file in os.listdir(company_folder):
                        file_path_to_remove = os.path.join(company_folder, file)
                        for attempt in range(max_retries):
                            try:
                                os.remove(file_path_to_remove)
                                files_removed += 1
                                print(f"Successfully removed {file_path_to_remove}")
                                break  # Exit retry loop on success
                            except PermissionError as e:
                                print(f"Attempt {attempt + 1} of {max_retries} failed to remove {file_path_to_remove}: {e}")
                                if attempt < max_retries - 1:
                                    time.sleep(retry_delay)
                                else:
                                    print(f"❌ Failed to remove {file_path_to_remove} after {max_retries} attempts.")
                            except Exception as e: # Catch other potential errors during removal
                                print(f"❌ An unexpected error occurred while trying to remove {file_path_to_remove}: {e}")
                                break # Exit retry loop for other errors
                    
                    #remove the company folder after uploading to blob storage
                    os.rmdir(company_folder)

                number_of_files_uploaded = len(urls_of_files_uploaded)

                print(f"uploaded {number_of_files_uploaded} files to blob storage /n")
                print(f"Removed {files_removed} files from the local folder /n")

        except requests.RequestException as e:
            failed_to_download_files.append(company_names[i])
            print(f"❌ Error downloading files for {company_names[i]}: {str(e)}")

    return company_names, xml_file_names, brsr_file_names, urls_of_files_uploaded, all_uploaded_blobs, failed_to_download_files
